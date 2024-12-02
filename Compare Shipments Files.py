import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the folder paths
folder_old_file = r'S:\Old File'
folder_new_file = r'S:\New File'

# Function to get the latest file from a directory (either .xlsx or .csv)
def get_latest_file_from_dir(directory):
    files = glob.glob(os.path.join(directory, "*.xlsx")) + glob.glob(os.path.join(directory, "*.csv"))
    if not files:
        raise FileNotFoundError(f"No files found in {directory}")
    return max(files, key=os.path.getmtime)


# Function to read a file (either .xlsx or .csv) and return it as a DataFrame
def read_file(file_path):
    if file_path.endswith('.csv'):
        return pd.read_csv(file_path)
    elif file_path.endswith('.xlsx'):
        return pd.read_excel(file_path)
    else:
        raise ValueError("Unsupported file format. Only .csv and .xlsx are supported.")


# Function to convert a DataFrame to Excel format
def convert_to_excel(df, output_path):
    df.to_excel(output_path, index=False)


try:
    # Step 1: Read the latest file from the old file folder
    old_file_path = get_latest_file_from_dir(folder_old_file)
    print(f"Reading Old File from: {old_file_path}")
    old_df = read_file(old_file_path)

    # Convert CSV to Excel if necessary
    if old_file_path.endswith('.csv'):
        old_file_path = old_file_path.replace('.csv', '.xlsx')
        convert_to_excel(old_df, old_file_path)
        print(f"Converted Old CSV to Excel: {old_file_path}")

except FileNotFoundError as e:
    print(e)
    exit(1)
except Exception as e:
    print(f"Error reading the old file: {e}")
    exit(1)

# Extract the old file name (without extension)
old_file_name = os.path.basename(old_file_path).replace('.xlsx', '')

try:
    # Step 2: Read the latest file from the new file folder
    new_file_path = get_latest_file_from_dir(folder_new_file)
    print(f"Reading New File from: {new_file_path}")
    new_df = read_file(new_file_path)

    # Convert CSV to Excel if necessary
    if new_file_path.endswith('.csv'):
        new_file_path = os.path.join(folder_new_file, f"{old_file_name}.xlsx")
        convert_to_excel(new_df, new_file_path)
        print(f"Converted New CSV to Excel: {new_file_path}")

except FileNotFoundError as e:
    print(e)
    exit(1)
except Exception as e:
    print(f"Error reading the new file: {e}")
    exit(1)

# Step 3: Ensure both required columns are present in the files
required_columns = ['Shipment name', 'Shipment ID']
missing_columns_old = [col for col in required_columns if col not in old_df.columns]
missing_columns_new = [col for col in required_columns if col not in new_df.columns]

if missing_columns_old:
    print(f"Columns {missing_columns_old} not found in the Old File.")
    exit(1)
if missing_columns_new:
    print(f"Columns {missing_columns_new} not found in the New File.")
    exit(1)

try:
    # Step 4: Merge the data to identify matching rows
    merged_df = pd.merge(new_df, old_df, on=['Shipment name', 'Shipment ID'], how='left', indicator=True)

    # Get the indices of matching rows in the new file
    matching_indices = merged_df[merged_df['_merge'] == 'both'].index

    # Step 5: Apply highlighting to the newly converted Excel file
    wb = load_workbook(new_file_path)
    ws = wb.active

    # Define the highlight color (yellow)
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply the highlight to matching rows
    for row_idx in matching_indices:
        excel_row = row_idx + 2  # Adjust for zero-indexing and header row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = highlight_fill

    # Step 6: Adjust column width to fit the content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (like 'A', 'B', etc.)

        # Find the maximum length of content in the column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        # Adjust the width with some padding
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the changes to the Excel file
    wb.save(new_file_path)
    print(f"Matching rows highlighted and saved to: {new_file_path}")

except Exception as e:
    print(f"Error during highlighting and saving: {e}")
    exit(1)
