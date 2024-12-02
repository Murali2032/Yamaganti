import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Path to the folder containing Excel files (including subfolders)
root_directory = r"C:\Users\YamagantiMuraliKrish\Music\November"
# Output Excel file to store results
output_file = r"C:\Users\YamagantiMuraliKrish\Music\November\Currency_Totals.xlsx"

# Exchange rates for dynamic conversion
exchange_rates = {
    "USD": 1.0,  # Base currency
    "INR": 0.0119,  # Indian Rupee to USD
    "CAD": 0.73,  # Canadian Dollar to USD
    "JPY": 0.0066,  # Japanese Yen to USD
    "EUR": 1.09,  # Euro to USD
    "AUD": 0.66,  # Australian Dollar to USD
    "GBP": 1.29  # British Pound to USD
}


def apply_formatting(output_path):
    """
    Apply formatting to the Excel file: bold headers, coloring, and set height and width.
    """
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb.active

    # Apply formatting to the header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark Gray
    alignment_center = Alignment(horizontal="center")
    alignment_left = Alignment(horizontal="left")
    alignment_right = Alignment(horizontal="right")

    # Set column width and row height
    for col in ws.columns:
        max_length = 0
        col_name = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_name].width = 42  # Set column width to 42

    # Apply header formatting
    for cell in ws[1]:  # First row is the header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        ws.row_dimensions[1].height = 20  # Set header row height to 20

    # Apply formatting to the body
    body_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light Gray

    for row in ws.iter_rows(min_row=2):  # From second row onwards (data rows)
        for cell in row:
            cell.fill = body_fill
            cell.alignment = alignment_center
        ws.row_dimensions[row[0].row].height = 20  # Set body row height to 20

    # Apply alignment for each column
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = alignment_left  # File Name column aligned to left
        row[1].alignment = alignment_right  # Total (USD) column aligned to right

    # Save the formatted workbook
    wb.save(output_path)
    print(f"Formatting applied to {output_path}")


def process_excel_files(directory, output_path):
    """
    Process all Excel files in the given directory and its subdirectories.
    Convert various currencies to USD, calculate the reimbursement totals,
    and save results into a new Excel file with formatting.
    """
    results = []  # To store file names and their totals

    # Traverse all files in the directory and subdirectories using os.walk
    for root, _, files in os.walk(directory):
        for filename in files:
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                file_path = os.path.join(root, filename)

                try:
                    # Read the Excel file
                    data = pd.read_excel(file_path)

                    # Check if required columns exist
                    if "Currency" in data.columns and "Reimbursement Total" in data.columns:
                        # Add a column for USD conversion
                        data["Reimbursement Total (USD)"] = data.apply(
                            lambda row: row["Reimbursement Total"] * exchange_rates.get(row["Currency"], 0),
                            axis=1
                        )

                        # Calculate the total in USD
                        total_usd = data["Reimbursement Total (USD)"].sum()

                        # Append the result
                        results.append({"File Name": filename, "Total (USD)": total_usd})
                        print(f"Processed {filename}: Total USD = {total_usd}")
                    else:
                        print(f"Skipped {filename}: Required columns missing.")
                except Exception as e:
                    print(f"Error processing {filename}: {e}")

    # Save results to a new Excel file
    if results:
        result_df = pd.DataFrame(results)
        result_df.to_excel(output_path, index=False)
        print(f"Results saved to {output_path}")
        # Apply formatting to the output file
        apply_formatting(output_path)
    else:
        print("No valid data to process.")


if __name__ == "__main__":
    print("Starting to process Excel files...")
    process_excel_files(root_directory, output_file)
    print("Processing completed!")
