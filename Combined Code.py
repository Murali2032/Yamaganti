import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Define the folder paths
input_folder_path = r'C:\Users\YamagantiMuraliKrish\OneDrive - Refuse Specialists LLC\Documents\New folder\Bank'
output_folder_path = r'C:\Users\YamagantiMuraliKrish\OneDrive - Refuse Specialists LLC\Documents\Final'

# Ensure the output folder exists; if not, create it
if not os.path.exists(output_folder_path):
    os.makedirs(output_folder_path)

# Get the latest Excel file from the input folder
files = [f for f in os.listdir(input_folder_path) if f.endswith('.xlsx')]
latest_file = max([os.path.join(input_folder_path, f) for f in files], key=os.path.getctime)

# Load the Excel file
full_df = pd.read_excel(latest_file, header=None)

# Separate the first 7 rows (header) and the rest of the data
header_rows = full_df.iloc[:7]  # First 7 rows
data_rows = full_df.iloc[7:]  # Rows from 8 onward

# Check if there are any rows left to process
if data_rows.shape[0] > 0:
    # Assuming all the relevant data is in the first column
    # Split the data in the first column by the pipe delimiter
    split_df = data_rows.iloc[:, 0].str.split('|', expand=True)

    # Create a list to hold all rows
    final_rows = []

    for i in range(split_df.shape[0]):
        row_data = split_df.iloc[i]
        # If all values in the row are NaN, retain the original value
        if row_data.isnull().all():
            final_rows.append({'Original': data_rows.iloc[i, 0]})
        else:
            # Append the split data to final rows
            final_rows.append(row_data.to_dict())

    # Convert the list of rows to a DataFrame
    processed_data_df = pd.DataFrame(final_rows)
else:
    processed_data_df = pd.DataFrame()  # Empty DataFrame if no data to process

# Combine the header rows with the processed data
final_df = pd.concat([header_rows, processed_data_df], ignore_index=True)

# Specify the columns to convert to numeric (C, D, E, F, G, H)
columns_to_convert = [2, 3, 4, 5, 6, 7]  # Using zero-based indexing for columns C-H

# Convert specified columns to numeric while keeping existing non-numeric values
for col in columns_to_convert:
    if col < final_df.shape[1]:  # Check if the column index is valid
        final_df[col] = final_df[col].apply(lambda x: pd.to_numeric(x, errors='coerce') if isinstance(x, str) and x.strip().replace('.', '', 1).isdigit() else x)

# Create a temporary file path to hold intermediate processed data
temp_output_excel_path = os.path.join(output_folder_path, "temp_output_bank_statement.xlsx")

# Save the final DataFrame to an Excel file without headers
final_df.to_excel(temp_output_excel_path, index=False, header=False)

# Load the workbook and select the active worksheet
wb = load_workbook(temp_output_excel_path)
ws = wb.active

# Align cells and set a fixed column width of 30 characters
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='top')  # Align left and top

# Set all columns to a width of 30
for column in ws.columns:
    ws.column_dimensions[column[0].column_letter].width = 30

# Ensure the number formatting in Excel for converted cells
for col in columns_to_convert:
    if col < final_df.shape[1]:  # Check if the column index is valid
        for row in range(8, ws.max_row + 1):  # Start from row 8
            cell = ws.cell(row=row, column=col + 1)  # +1 because column index is 1-based in Excel
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'  # Set the format to number with two decimal places

# Save the workbook to the temporary file
wb.save(temp_output_excel_path)

# Now proceed to split the Excel file based on the separator and format the output
def split_excel_file(input_file, output_file, separator="____________________________________________________________________________________________________________________________________"):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(input_file, header=None)

    # Initialize variables
    sheet_num = 1
    start_index = 0

    # Create a dictionary to hold data for each sheet
    sheets_data = {}

    # Iterate through the DataFrame and find rows that contain the separator
    for i, row in df.iterrows():
        # If the separator is found
        if separator in str(row.values[0]):
            # Slice the DataFrame and add the data to the dictionary
            sheet_data = df.iloc[start_index:i]
            if not sheet_data.empty:
                sheets_data[f'Sheet{sheet_num}'] = sheet_data

            # Update indices for the next sheet
            start_index = i + 1
            sheet_num += 1

    # Add remaining data to the last sheet (if any)
    if start_index < len(df):
        sheets_data[f'Sheet{sheet_num}'] = df.iloc[start_index:]

    # Write the data to a new Excel file with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, sheet_data in sheets_data.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # After saving, reopen the workbook to apply formatting (alignment and column width)
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        # Align cells and set a fixed column width of 30 characters
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top')  # Align left and top

        # Set all columns to a width of 30
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

    # Save the changes to the file
    wb.save(output_file)
    print(f"Data has been successfully split and formatted in '{output_file}'.")

# Final output file path
final_output_excel_path = os.path.join(output_folder_path, "final_output_bank_statement.xlsx")

# Call the function to split the Excel file and apply formatting to the final output
split_excel_file(temp_output_excel_path, final_output_excel_path)

# Remove the temporary file after processing is done
if os.path.exists(temp_output_excel_path):
    os.remove(temp_output_excel_path)

print(f"Final formatted and split Excel file created successfully at: {final_output_excel_path}")
