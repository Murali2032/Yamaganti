import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the folder paths
folder_old_file = r'S:\Old File'
folder_new_file = r'S:\New File'

# Function to find the first Excel file in a folder
def find_excel_file(folder_path):
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            return os.path.join(folder_path, file_name)
    return None

# Get the Excel files from both folders
file_old = find_excel_file(folder_old_file)
file_new = find_excel_file(folder_new_file)

if file_old and file_new:
    # Load the Excel sheets into dataframes
    df_old = pd.read_excel(file_old)
    df_new = pd.read_excel(file_new)

    # Remove duplicates in each dataframe (if necessary)
    df_old = df_old.drop_duplicates()
    df_new = df_new.drop_duplicates()

    # Finding duplicates in 'Reimbursement ID', 'Claims Number', and 'Date'
    duplicate_reimbursement = pd.merge(df_old[['Reimbursement ID']], df_new[['Reimbursement ID']], on='Reimbursement ID', how='inner')
    duplicate_claims = pd.merge(df_old[['Claims Number']], df_new[['Claims Number']], on='Claims Number', how='inner')
    duplicate_date = pd.merge(df_old[['Date']], df_new[['Date']], on='Date', how='inner')

    # Load the New File Excel into openpyxl workbook
    wb = load_workbook(file_new)
    ws = wb.active

    # Print headers to verify column positions
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    print(f"Headers in the worksheet: {headers}")

    # Verify column index of 'Reimbursement ID', 'Claims Number', and 'Date'
    reimbursement_col = headers.index('Reimbursement ID') + 1  # Adding 1 since index is 0-based
    claims_col = headers.index('Claims Number') + 1
    date_col = headers.index('Date') + 1

    print(f"Reimbursement ID is in column {reimbursement_col}")
    print(f"Claims Number is in column {claims_col}")
    print(f"Date is in column {date_col}")

    # Define fill colors for highlighting duplicates
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight

    # Highlight duplicates in 'Reimbursement ID'
    if not duplicate_reimbursement.empty:
        print("Highlighting Reimbursement IDs:")
        for reimbursement_id in duplicate_reimbursement['Reimbursement ID']:
            for row in range(2, ws.max_row + 1):  # Assuming the first row is the header
                if ws.cell(row=row, column=reimbursement_col).value == reimbursement_id:
                    ws.cell(row=row, column=reimbursement_col).fill = highlight_fill
                    print(f"Highlighted Reimbursement ID: {reimbursement_id} at row {row}")

    # Highlight duplicates in 'Claims Number'
    if not duplicate_claims.empty:
        print("Highlighting Claims Numbers:")
        for claims_number in duplicate_claims['Claims Number']:
            for row in range(2, ws.max_row + 1):  # Assuming the first row is the header
                if ws.cell(row=row, column=claims_col).value == claims_number:
                    ws.cell(row=row, column=claims_col).fill = highlight_fill
                    print(f"Highlighted Claims Number: {claims_number} at row {row}")

    # Highlight entire row for duplicate 'Date'
    if not duplicate_date.empty:
        print("Highlighting entire rows for duplicate Dates:")
        for date_value in duplicate_date['Date']:
            for row in range(2, ws.max_row + 1):  # Assuming the first row is the header
                if ws.cell(row=row, column=date_col).value == date_value:
                    for col in range(1, ws.max_column + 1):  # Highlight the entire row
                        ws.cell(row=row, column=col).fill = highlight_fill
                    print(f"Highlighted entire row {row} for Date: {date_value}")

    # Save changes to the same file
    wb.save(file_new)
    print(f"Duplicates highlighted and saved in the same file: {file_new}")
else:
    print("No Excel file found in one of the folders.")
