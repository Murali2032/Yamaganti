import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def split_excel_file(input_file, output_file, separator="____________________________________________________________________________________________________________________________________"):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(input_file, header=None)

    # Initialize variables
    sheet_num = 1
    start_index = 0

    # Create a dictionary to hold data for each sheet
    sheets_data = {}

    # Iterate through the DataFrame and find rows that contain the separator
    for i, row in df.iterrows():
        # If the separator is found
        if separator in str(row.values[0]):
            # Slice the DataFrame and add the data to the dictionary
            sheet_data = df.iloc[start_index:i]
            sheets_data[f'Sheet{sheet_num}'] = sheet_data

            # Update indices for the next sheet
            start_index = i + 1
            sheet_num += 1

    # Add remaining data to the last sheet
    if start_index < len(df):
        sheets_data[f'Sheet{sheet_num}'] = df.iloc[start_index:]

    # Write the data to a new Excel file with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, sheet_data in sheets_data.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # After saving, reopen the workbook to apply formatting (alignment and column width)
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        # Align cells and set a fixed column width of 30 characters
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top')  # Align left and top

        # Set all columns to a width of 30
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

    # Save the changes to the file
    wb.save(output_file)
    print(f"Data has been successfully split and formatted in '{output_file}'.")

def process_folder(input_folder_path, output_folder_path):
    # Check if output folder exists, if not, create it
    if not os.path.exists(output_folder_path):
        os.makedirs(output_folder_path)

    # Iterate through all Excel files in the input folder
    for filename in os.listdir(input_folder_path):
        if filename.endswith(".xlsx"):  # Process only Excel files
            input_file = os.path.join(input_folder_path, filename)
            output_file = os.path.join(output_folder_path, f"split_{filename}")

            # Call the function to split the Excel file and apply formatting
            split_excel_file(input_file, output_file)

# Example usage
input_folder_path = r'C:\Users\YamagantiMuraliKrish\Documents\New folder\Bank'
output_folder_path = r'C:\Users\YamagantiMuraliKrish\Documents\Output'

process_folder(input_folder_path, output_folder_path)
