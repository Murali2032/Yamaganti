import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the directory paths
input_directory = r'S:\\New File'
output_directory = r'S:\\New File\\Output'  # Folder to save the Excel file
folder_old_file = r'S:\\Old File'  # Path to the old Excel file

# Function to get the latest CSV file from the directory
def get_latest_csv_file(directory):
    csv_files = [f for f in os.listdir(directory) if f.endswith('.csv')]
    if not csv_files:
        raise FileNotFoundError("No CSV files found in the directory.")
    csv_files.sort(key=lambda f: os.path.getmtime(os.path.join(directory, f)), reverse=True)
    return os.path.join(directory, csv_files[0])

try:
    # Get the latest CSV file
    input_file = get_latest_csv_file(input_directory)

    # Read the CSV file into a pandas DataFrame with the correct encoding
    df = pd.read_csv(input_file, encoding='ISO-8859-1')

    # Debugging: Print the columns in the CSV file
    print("Columns in the CSV file:", df.columns)

    # Check if the number of columns is greater than the specified indices to drop
    columns_to_drop_indices = [3, 5, 7, 8, 9, 13, 14, 15, 16]
    valid_columns_to_drop = [i for i in columns_to_drop_indices if i < len(df.columns)]

    # Drop valid columns
    df.drop(df.columns[valid_columns_to_drop], axis=1, inplace=True)

    # Update the column names
    column_mapping = {
        'approval-date': 'Date',
        'reimbursement-id': 'Reimbursement ID',
        'case-id': 'Claims Number',
        'reason': 'Reimbursement Type',  # Adjust for case sensitivity
        'fnsku': 'FNSKU',
        'currency-unit': 'Currency',
        'per-unit-reimbursement-amount': 'Amount Per Unit',
        'amount-total': 'Reimbursement Total',  # Update "amount-total" to "Reimbursement Total"
        'total-reimbursement': 'Reimbursement Total',  # Also handle "total-reimbursement"
    }
    df.rename(columns=column_mapping, inplace=True)

    # Drop the "original-reimbursement-type" column after renaming
    df.drop(columns=["original-reimbursement-type"], errors='ignore', inplace=True)

    # Debugging: Print columns after renaming
    print("Columns after renaming:", df.columns)

    # Ensure the output directory exists
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    # Reading the old Excel file from folder_old_file
    old_file = None
    for file_name in os.listdir(folder_old_file):
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            old_file = os.path.join(folder_old_file, file_name)
            break

    if old_file:
        # Get the base name of the old file (without extension)
        old_file_name = os.path.basename(old_file).split('.')[0]

        # Remove any existing month name from the file name
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        for month in months:
            if month in old_file_name:
                old_file_name = old_file_name.replace(month, "")

        # Get the current month name
        current_month = "November"  # Replace with the desired month

        # Construct the new output file name
        output_file_name = f"{old_file_name.strip()}-{current_month}.xlsx"

        # Capitalize the first letter of each word in the file name
        output_file_name = ' '.join([word.capitalize() for word in output_file_name.split('-')])

        # Save the modified DataFrame to an Excel file with the new name
        output_file = os.path.join(output_directory, output_file_name)
        df.to_excel(output_file, index=False)

        # Load the Excel file for further formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define the header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_alignment = Alignment(horizontal='left', vertical='center')  # Left-aligned, vertically centered

        # Apply header styles and adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)  # Add a little padding to the width
            ws.column_dimensions[col[0].column_letter].width = adjusted_width  # Adjust column width

        # Set header row height to 22
        ws.row_dimensions[1].height = 22

        # Apply the styles to the header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Define a color for highlighting the "0" values
        zero_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color

        # Highlight "0" values in the "Reimbursement Total" column
        reimbursement_total_column_index = df.columns.get_loc('Reimbursement Total') + 1  # Get the column index (1-based)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=reimbursement_total_column_index,
                                max_col=reimbursement_total_column_index):
            for cell in row:
                if cell.value == 0:
                    cell.fill = zero_fill  # Highlight cells with value 0

        # Save the updated Excel file with formatting
        wb.save(output_file)

        print(f"CSV has been successfully converted to Excel, formatted, and saved as '{output_file}'.")

        # Load the old Excel file to compare with the new one
        df_old = pd.read_excel(old_file)

        # Define the fill color for highlighting duplicates
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow color

        # Compare each row between the new file and the old file based on the 2 key columns
        for new_row_idx, new_row in df.iterrows():
            new_row_values = (new_row['Date'], new_row['Reimbursement ID'])
            matching_rows = df_old[
                (df_old['Date'] == new_row['Date']) & (df_old['Reimbursement ID'] == new_row['Reimbursement ID'])
            ]

            if not matching_rows.empty:
                # Highlight the entire row in the new Excel file if a match is found
                for row in ws.iter_rows(min_row=new_row_idx + 2, max_row=new_row_idx + 2, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.fill = highlight_fill  # Apply yellow highlight

        # Save the final updated Excel file with highlighted duplicates
        wb.save(output_file)
        print(f"Duplicates based on 'Date' and 'Reimbursement ID' between the new and old files have been highlighted in '{output_file}'.")

except Exception as e:
    print(f"An error occurred: {e}")
